import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from app.models.product import Product
from app.models.category import Category
from app.models.temper_glass import TemperBox


def style_header(ws, headers, header_color="1565C0"):
    """Apply professional styling to header row"""
    fill = PatternFill(start_color=header_color,
                       end_color=header_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = alignment
        ws.column_dimensions[
            cell.column_letter
        ].width = max(len(header) + 8, 15)

    ws.row_dimensions[1].height = 30


def style_data_row(ws, row_num, num_cols, is_low_stock=False):
    """Style data rows with alternating colors"""
    fill_color = "FFF3CD" if is_low_stock else (
        "F8F9FA" if row_num % 2 == 0 else "FFFFFF"
    )
    fill = PatternFill(start_color=fill_color,
                       end_color=fill_color, fill_type="solid")
    border = Border(
        left=Side(style='thin', color="DEE2E6"),
        right=Side(style='thin', color="DEE2E6"),
        top=Side(style='thin', color="DEE2E6"),
        bottom=Side(style='thin', color="DEE2E6")
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)


# ─── MOBILE COVERS ────────────────────────────────────────────────────────────
def generate_mobile_covers_excel(category_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mobile Covers"

    headers = ["#", "Brand", "Mobile Model",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    products = Product.query.filter_by(
        category_id=category_id, is_active=True
    ).all()

    for i, p in enumerate(products, 1):
        status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
        row = [
            i, p.brand, p.mobile_model,
            p.quantity, status,
            p.notes or '',
            p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
        ]
        ws.append(row)
        style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    return _save_workbook(wb)


# ─── EARPHONES ────────────────────────────────────────────────────────────────
def generate_earphones_excel(category_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Earphones"

    headers = ["#", "Brand", "Model",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    products = Product.query.filter_by(
        category_id=category_id, is_active=True
    ).all()

    for i, p in enumerate(products, 1):
        status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
        row = [
            i, p.brand, p.product_name,
            p.quantity, status,
            p.notes or '',
            p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
        ]
        ws.append(row)
        style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    return _save_workbook(wb)


# ─── CHARGERS ─────────────────────────────────────────────────────────────────
def generate_chargers_excel(category_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chargers"

    headers = ["#", "Brand", "Watts",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    products = Product.query.filter_by(
        category_id=category_id, is_active=True
    ).all()

    for i, p in enumerate(products, 1):
        status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
        row = [
            i, p.brand, p.watts,
            p.quantity, status,
            p.notes or '',
            p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
        ]
        ws.append(row)
        style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    return _save_workbook(wb)


# ─── CHARGER CABLES ───────────────────────────────────────────────────────────
def generate_cables_excel(category_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Charger Cables"

    headers = ["#", "Brand", "Cable Type",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    products = Product.query.filter_by(
        category_id=category_id, is_active=True
    ).all()

    for i, p in enumerate(products, 1):
        status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
        row = [
            i, p.brand, p.cable_type,
            p.quantity, status,
            p.notes or '',
            p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
        ]
        ws.append(row)
        style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    return _save_workbook(wb)


# ─── TEMPER GLASS ─────────────────────────────────────────────────────────────
def generate_temper_glass_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Temper Glass"

    headers = ["#", "Box Name", "Mobile Model",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    boxes = TemperBox.query.filter_by(is_active=True).all()
    row_num = 2
    counter = 1

    for box in boxes:
        for item in box.items:
            status = "⚠ Low Stock" if item.quantity < 3 else "✓ In Stock"
            row = [
                counter, box.box_name, item.mobile_model,
                item.quantity, status,
                item.notes or '',
                item.created_at.strftime('%d-%m-%Y') if item.created_at else ''
            ]
            ws.append(row)
            style_data_row(ws, row_num, len(headers), item.quantity < 3)
            row_num += 1
            counter += 1

    return _save_workbook(wb)


# ─── OTHERS / CUSTOM CATEGORIES ───────────────────────────────────────────────
def generate_others_excel(category_id):
    wb = Workbook()
    ws = wb.active

    category = Category.query.get(category_id)
    ws.title = category.name[:31] if category else "Others"

    headers = ["#", "Brand", "Product Name",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers)

    products = Product.query.filter_by(
        category_id=category_id, is_active=True
    ).all()

    for i, p in enumerate(products, 1):
        status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
        row = [
            i, p.brand, p.product_name,
            p.quantity, status,
            p.notes or '',
            p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
        ]
        ws.append(row)
        style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    return _save_workbook(wb)


# ─── FULL INVENTORY (ALL CATEGORIES) ─────────────────────────────────────────
def generate_full_inventory_excel():
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    categories = Category.query.filter_by(is_active=True).all()

    for category in categories:
        ws = wb.create_sheet(title=category.name[:31])
        headers = ["#", "Brand", "Product/Model",
                   "Extra Info", "Quantity", "Stock Status", "Notes", "Added Date"]
        style_header(ws, headers, "1565C0")

        products = Product.query.filter_by(
            category_id=category.id, is_active=True
        ).all()

        for i, p in enumerate(products, 1):
            extra = p.watts or p.cable_type or p.mobile_model or ''
            status = "⚠ Low Stock" if p.quantity < 3 else "✓ In Stock"
            row = [
                i,
                p.brand or '',
                p.product_name or p.mobile_model or '',
                extra,
                p.quantity,
                status,
                p.notes or '',
                p.created_at.strftime('%d-%m-%Y') if p.created_at else ''
            ]
            ws.append(row)
            style_data_row(ws, i + 1, len(headers), p.quantity < 3)

    # Temper Glass sheet
    ws = wb.create_sheet(title="Temper Glass")
    headers = ["#", "Box Name", "Mobile Model",
               "Quantity", "Stock Status", "Notes", "Added Date"]
    style_header(ws, headers, "1565C0")

    boxes = TemperBox.query.filter_by(is_active=True).all()
    row_num = 2
    counter = 1
    for box in boxes:
        for item in box.items:
            status = "⚠ Low Stock" if item.quantity < 3 else "✓ In Stock"
            ws.append([
                counter, box.box_name, item.mobile_model,
                item.quantity, status,
                item.notes or '',
                item.created_at.strftime('%d-%m-%Y') if item.created_at else ''
            ])
            style_data_row(ws, row_num, len(headers), item.quantity < 3)
            row_num += 1
            counter += 1

    return _save_workbook(wb)


def _save_workbook(wb):
    """Save workbook to bytes buffer"""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer